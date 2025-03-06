from PyPDF2 import PdfReader
from docx import Document
import openpyxl
from io import BytesIO
import os
import io

import pytesseract
from PIL import Image

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

BIN_EXT = ['pdf', 'docx', 'xlsx', 'xls', 'png', 'jpg', 'jpeg']
TEXT_EXT = ['txt', 'java', 'csv', 'py', 'html', 'css', 'js', 'json', 'log']
ALLOWED_EXTENSIONS = TEXT_EXT + BIN_EXT

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def read_txt(file):
    return file.read().decode('utf-8')

def read_image(file):
    try: 
        image = Image.open(io.BytesIO(file.read()))
        return pytesseract.image_to_string(image)
    except pytesseract.pytesseract.TesseractNotFoundError as e:
        error_text = r"""
            Error reading Image File.
            Please make sure you have installed Tesseract-OCR and added it to your system's PATH.
            Install Tesseract-OCR from https://github.com/tesseract-ocr/tesseract/releases.
            Also check that Tesseract-OCR is installed in "C:\Program Files\Tesseract-OCR\tesseract.exe"
            if not update the same path in line 10.
        """
        print(error_text)
        return error_text

def read_pdf(file):
    pdf_reader = PdfReader(BytesIO(file.read()))
    return "".join([page.extract_text() for page in pdf_reader.pages])

def read_docx(file):
    doc = Document(BytesIO(file.read()))
    return "\n".join([para.text for para in doc.paragraphs])

def read_xlsx(file):
    file_content = "";
    wb = openpyxl.load_workbook(file)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        file_content += f", sheet {sheet_name} :["
        file_content += "\n".join([", ".join([sheet.cell(row=r, column=c).value
                                                    for c in range(1, sheet.max_column + 1)]
                                                        ) for r in range(1, sheet.max_row + 1)])
        file_content += "]\n\n"
    return file_content
        
def read_file(file):
    
    _, ext = os.path.splitext(file.filename)
    print(ext)
    if ext == '.png' or ext == '.jpg' or ext == '.jpeg':
        return read_image(file)
    if ext == '.pdf':
        return read_pdf(file)
    elif ext == '.docx':
        return read_docx(file)
    elif ext[1:] in TEXT_EXT:
        return read_txt(file)
    elif ext == '.xlsx' or ext == '.xls':
        return read_xlsx(file)